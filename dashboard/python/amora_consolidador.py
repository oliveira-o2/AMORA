import argparse
import json
import logging
from numbers import Number
import re
import sys
import unicodedata
import unittest
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

try:
    import tkinter as tk  # type: ignore
    from tkinter import filedialog, messagebox  # type: ignore
    TKINTER_AVAILABLE = True
except ModuleNotFoundError:
    tk = None
    filedialog = None
    messagebox = None
    TKINTER_AVAILABLE = False


# =========================================================
# CONFIGURAÇÃO GERAL
# =========================================================
APP_NAME = "Dash Amora - Consolidador"
DEFAULT_OUTPUT_PREFIX = "Consolidado_Amora"
LOG_FILENAME = "amora_consolidador.log"


# =========================================================
# LOGGING
# =========================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILENAME, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)


# =========================================================
# ESTRUTURAS DE DADOS
# =========================================================
@dataclass
class ProcessingResult:
    consolidado: pd.DataFrame
    guia_cfop: pd.DataFrame
    inconsistencias: pd.DataFrame
    metricas: pd.DataFrame
    caminho_saida: Path
    caminho_json: Path


@dataclass
class UserInputs:
    arquivos_nfe: List[str]
    caminho_vendas_item: Optional[str]
    caminho_lista_vendas: Optional[str]
    caminho_saida: Path


# =========================================================
# ABSTRAÇÃO DE INTERFACE
# =========================================================
class UserInterface:
    def info(self, titulo: str, mensagem: str) -> None:
        raise NotImplementedError

    def warning(self, titulo: str, mensagem: str) -> None:
        raise NotImplementedError

    def error(self, titulo: str, mensagem: str) -> None:
        raise NotImplementedError

    def ask_open_files(self, titulo: str) -> List[str]:
        raise NotImplementedError

    def ask_open_file(self, titulo: str) -> Optional[str]:
        raise NotImplementedError

    def ask_save_file(self, titulo: str, default_name: str) -> Optional[Path]:
        raise NotImplementedError

    def close(self) -> None:
        return None


class TkUserInterface(UserInterface):
    def __init__(self):
        if not TKINTER_AVAILABLE:
            raise RuntimeError("Tkinter não está disponível neste ambiente.")
        self.root = tk.Tk()
        self.root.withdraw()

    def info(self, titulo: str, mensagem: str) -> None:
        messagebox.showinfo(titulo, mensagem)

    def warning(self, titulo: str, mensagem: str) -> None:
        messagebox.showwarning(titulo, mensagem)

    def error(self, titulo: str, mensagem: str) -> None:
        messagebox.showerror(titulo, mensagem)

    def ask_open_files(self, titulo: str) -> List[str]:
        arquivos = filedialog.askopenfilenames(
            title=titulo,
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")],
        )
        return list(arquivos)

    def ask_open_file(self, titulo: str) -> Optional[str]:
        arquivo = filedialog.askopenfilename(
            title=titulo,
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")],
        )
        return arquivo or None

    def ask_save_file(self, titulo: str, default_name: str) -> Optional[Path]:
        caminho = filedialog.asksaveasfilename(
            title=titulo,
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Arquivo Excel", "*.xlsx")],
        )
        return Path(caminho) if caminho else None

    def close(self) -> None:
        try:
            self.root.destroy()
        except Exception:
            pass


class ConsoleUserInterface(UserInterface):
    def info(self, titulo: str, mensagem: str) -> None:
        print(f"[INFO] {titulo}: {mensagem}")

    def warning(self, titulo: str, mensagem: str) -> None:
        print(f"[AVISO] {titulo}: {mensagem}")

    def error(self, titulo: str, mensagem: str) -> None:
        print(f"[ERRO] {titulo}: {mensagem}", file=sys.stderr)

    def ask_open_files(self, titulo: str) -> List[str]:
        print(f"{titulo}")
        print("Informe os caminhos dos arquivos Excel separados por ponto e vírgula (;):")
        entrada = input().strip()
        if not entrada:
            return []
        return [item.strip() for item in entrada.split(";") if item.strip()]

    def ask_open_file(self, titulo: str) -> Optional[str]:
        print(f"{titulo}")
        print("Informe o caminho do arquivo Excel ou pressione ENTER para pular:")
        entrada = input().strip()
        return entrada or None

    def ask_save_file(self, titulo: str, default_name: str) -> Optional[Path]:
        print(f"{titulo}")
        print(f"Informe o caminho de saída ou pressione ENTER para usar {default_name} no diretório atual:")
        entrada = input().strip()
        if not entrada:
            return Path(default_name)
        return Path(entrada)


# =========================================================
# FUNÇÕES UTILITÁRIAS
# =========================================================
def build_default_output_name() -> str:
    return f"{DEFAULT_OUTPUT_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def normalizar_texto(valor: str) -> str:
    if valor is None:
        return ""
    valor = str(valor).strip()
    valor = valor.replace("&", " e ")
    valor = unicodedata.normalize("NFKD", valor).encode("ASCII", "ignore").decode("ASCII")
    valor = re.sub(r"\s+", " ", valor)
    return valor.lower().strip()


def normalizar_nome_coluna(coluna: str) -> str:
    coluna = normalizar_texto(coluna)
    coluna = coluna.replace(".", " ").replace("/", " ").replace("-", " ")
    coluna = re.sub(r"\s+", "_", coluna)
    return coluna


def padronizar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    novo_df = df.copy()
    novo_df.columns = [normalizar_nome_coluna(col) for col in novo_df.columns]
    return novo_df.loc[:, ~novo_df.columns.duplicated()].copy()


def ler_excel(caminho_arquivo: str) -> Optional[pd.DataFrame]:
    try:
        engine = "xlrd" if caminho_arquivo.lower().endswith(".xls") else "openpyxl"
        df = pd.read_excel(caminho_arquivo, engine=engine)
        logger.info(
            "Arquivo lido com sucesso: %s | linhas=%s colunas=%s",
            caminho_arquivo,
            df.shape[0],
            df.shape[1],
        )
        return df
    except Exception as e:
        logger.exception("Erro ao ler arquivo %s: %s", caminho_arquivo, e)
        return None


def converter_numero_br(valor):
    if pd.isna(valor):
        return pd.NA
    if isinstance(valor, Number) and not isinstance(valor, bool):
        return pd.to_numeric(valor, errors="coerce")
    texto = str(valor).strip()
    if texto == "":
        return pd.NA
    texto = texto.replace("R$", "").replace(" ", "")
    texto = texto.replace(".", "").replace(",", ".")
    return pd.to_numeric(texto, errors="coerce")


def normalizar_chave_item(valor) -> str:
    if pd.isna(valor):
        return ""
    if isinstance(valor, Number) and not isinstance(valor, bool):
        valor_float = float(valor)
        if valor_float.is_integer():
            return str(int(valor_float))
        return str(valor_float)
    texto = normalizar_texto(valor)
    return re.sub(r"\.0+$", "", texto)


def converter_coluna_numerica(df: pd.DataFrame, coluna: str) -> pd.Series:
    return df[coluna].apply(converter_numero_br)


def encontrar_coluna(df: pd.DataFrame, candidatos: List[str], obrigatoria: bool = False) -> Optional[str]:
    colunas = list(df.columns)
    for candidato in candidatos:
        candidato_norm = normalizar_nome_coluna(candidato)
        if candidato_norm in colunas:
            return candidato_norm
    if obrigatoria:
        raise KeyError(f"Nenhuma das colunas obrigatórias foi encontrada: {candidatos}")
    return None


def extrair_num_nf(valor) -> Optional[str]:
    if pd.isna(valor) or str(valor).strip() == "":
        return None
    if isinstance(valor, (int, float)) and not pd.isna(valor):
        try:
            valor_float = float(valor)
            if valor_float.is_integer():
                return str(int(valor_float))
        except (TypeError, ValueError):
            pass
    texto = str(valor).strip()
    if "/" in texto:
        texto = texto.split("/")[0]
    texto = re.sub(r"\.0+$", "", texto)
    texto = re.sub(r"nfs?e?", "", texto, flags=re.IGNORECASE).strip()
    numeros = re.sub(r"[^0-9]", "", texto)
    return numeros if numeros else None


def limpar_linhas_totais(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    primeira_coluna = df.columns[0]
    mask_total = ~df[primeira_coluna].astype(str).str.contains("total|soma|somatorio", case=False, na=False)
    df = df[mask_total].copy()
    if len(df.columns) > 2:
        df = df.dropna(subset=df.columns[2:], how="all")
    return df


def obter_interface(force_console: bool = False) -> UserInterface:
    if not force_console and TKINTER_AVAILABLE:
        logger.info("Interface gráfica Tkinter disponível. Usando modo GUI.")
        return TkUserInterface()
    logger.info("Tkinter indisponível ou modo console forçado. Usando modo console.")
    return ConsoleUserInterface()


def coletar_inputs(args: argparse.Namespace, ui: UserInterface) -> Optional[UserInputs]:
    arquivos_nfe = sorted(list(args.nfe or []))
    caminho_vendas_item = args.vendas_item
    caminho_lista_vendas = args.lista_vendas
    caminho_saida = Path(args.output) if args.output else None

    if not arquivos_nfe:
        ui.info(APP_NAME, "Passo 1: selecione os relatórios de Itens por NFe para consolidar.")
        arquivos_nfe = ui.ask_open_files("Selecione os relatórios de NFe")
    if not arquivos_nfe:
        ui.warning(APP_NAME, "Processo cancelado. Nenhum relatório de NFe foi selecionado.")
        return None
    arquivos_nfe = sorted(arquivos_nfe)

    if caminho_vendas_item is None and not args.no_prompt:
        ui.info(APP_NAME, "Passo 2: selecione o relatório de Vendas por Item para calcular o Custo Médio. Cancele ou deixe em branco para pular.")
        caminho_vendas_item = ui.ask_open_file("Selecione o relatório de Vendas por Item")

    if caminho_lista_vendas is None and not args.no_prompt:
        ui.info(APP_NAME, "Passo 3: selecione a Lista de Vendas para cruzar o Cst. Total por NF. Cancele ou deixe em branco para pular.")
        caminho_lista_vendas = ui.ask_open_file("Selecione a Lista de Vendas")

    if caminho_saida is None:
        caminho_saida = ui.ask_save_file("Salvar arquivo consolidado", build_default_output_name())
    if caminho_saida is None:
        ui.warning(APP_NAME, "Processo cancelado. Nenhum arquivo de saída foi definido.")
        return None

    return UserInputs(
        arquivos_nfe=arquivos_nfe,
        caminho_vendas_item=caminho_vendas_item,
        caminho_lista_vendas=caminho_lista_vendas,
        caminho_saida=caminho_saida,
    )


# =========================================================
# REGRAS DE NEGÓCIO
# =========================================================
def carregar_relatorios_nfe(arquivos_nfe: Sequence[str]) -> Tuple[pd.DataFrame, List[Dict[str, str]]]:
    lista_df = []
    inconsistencias: List[Dict[str, str]] = []

    for arquivo in arquivos_nfe:
        df = ler_excel(arquivo)
        if df is None or df.empty:
            inconsistencias.append(
                {
                    "etapa": "carregar_relatorios_nfe",
                    "arquivo": str(arquivo),
                    "tipo": "arquivo_invalido",
                    "detalhe": "Arquivo não pôde ser lido ou está vazio.",
                }
            )
            continue

        df = limpar_linhas_totais(df)
        df = padronizar_colunas(df)

        if df.empty:
            inconsistencias.append(
                {
                    "etapa": "carregar_relatorios_nfe",
                    "arquivo": str(arquivo),
                    "tipo": "sem_dados",
                    "detalhe": "Arquivo ficou sem linhas válidas após limpeza.",
                }
            )
            continue

        lista_df.append(df)

    if not lista_df:
        raise ValueError("Nenhum relatório de NFe válido foi encontrado.")

    colunas_base = lista_df[0].columns.tolist()
    lista_ajustada = []

    for idx, df in enumerate(lista_df):
        for col in colunas_base:
            if col not in df.columns:
                df[col] = pd.NA
        df = df[colunas_base]
        lista_ajustada.append(df)
        logger.info("Relatório NFe ajustado: %s", arquivos_nfe[idx])

    consolidado = pd.concat(lista_ajustada, ignore_index=True)

    col_data = consolidado.columns[0]
    consolidado[col_data] = pd.to_datetime(consolidado[col_data], errors="coerce", dayfirst=True).dt.date

    invalid_dates = consolidado[col_data].isna().sum()
    if invalid_dates:
        inconsistencias.append(
            {
                "etapa": "carregar_relatorios_nfe",
                "arquivo": "consolidado",
                "tipo": "datas_invalidas",
                "detalhe": f"{invalid_dates} linhas com data inválida na coluna {col_data}.",
            }
        )

    return consolidado, inconsistencias


def calcular_custo_medio(
    consolidado: pd.DataFrame, caminho_vendas_item: str
) -> Tuple[pd.DataFrame, List[Dict[str, str]], pd.DataFrame]:
    inconsistencias: List[Dict[str, str]] = []
    df_vendas = ler_excel(caminho_vendas_item)
    if df_vendas is None or df_vendas.empty:
        raise ValueError("Relatório de vendas por item inválido.")

    df_vendas = padronizar_colunas(df_vendas)
    consolidado = consolidado.copy()

    col_custo_total = encontrar_coluna(df_vendas, ["Custo total", "Cst. total", "Custo t", "Custo"])
    col_quantidade = encontrar_coluna(df_vendas, ["Quantidade", "Qtde", "Qtd"], obrigatoria=True)
    col_codigo_v = encontrar_coluna(df_vendas, ["Código", "Cod", "SKU", "Item"])
    col_nome_v = encontrar_coluna(df_vendas, ["Nome", "Descrição", "Produto", "Item"])

    if not col_custo_total:
        raise KeyError("Coluna de custo total não encontrada no relatório de vendas por item.")

    df_vendas["custo_total_num"] = converter_coluna_numerica(df_vendas, col_custo_total)
    df_vendas["quantidade_num"] = converter_coluna_numerica(df_vendas, col_quantidade)
    df_vendas["custo_medio"] = df_vendas["custo_total_num"] / df_vendas["quantidade_num"]

    if col_codigo_v:
        df_vendas["chave_cod"] = df_vendas[col_codigo_v].map(normalizar_chave_item)
    if col_nome_v:
        df_vendas["chave_nome"] = df_vendas[col_nome_v].astype(str).map(normalizar_texto)

    col_codigo_c = encontrar_coluna(consolidado, ["Código", "Cod", "SKU", "Item"])
    col_nome_c = encontrar_coluna(consolidado, ["Nome", "Descrição", "Produto", "Item"])
    col_qtd_c = encontrar_coluna(consolidado, ["Quantidade", "Qtde", "Qtd", "Quant"], obrigatoria=True)

    if col_codigo_c:
        consolidado["chave_cod"] = consolidado[col_codigo_c].map(normalizar_chave_item)
    if col_nome_c:
        consolidado["chave_nome"] = consolidado[col_nome_c].astype(str).map(normalizar_texto)

    consolidado["custo_medio_x"] = pd.NA
    consolidado["origem_match_custo"] = pd.NA

    if "chave_cod" in consolidado.columns and "chave_cod" in df_vendas.columns:
        mapa_cod = (
            df_vendas.dropna(subset=["chave_cod", "custo_medio"])
            .drop_duplicates(subset=["chave_cod"], keep="first")
            .set_index("chave_cod")["custo_medio"]
        )
        consolidado["custo_medio_x"] = consolidado["chave_cod"].map(mapa_cod)
        consolidado.loc[consolidado["custo_medio_x"].notna(), "origem_match_custo"] = "codigo"

        duplicados_cod = df_vendas["chave_cod"].duplicated(keep=False).sum()
        if duplicados_cod:
            inconsistencias.append(
                {
                    "etapa": "calcular_custo_medio",
                    "arquivo": caminho_vendas_item,
                    "tipo": "codigos_duplicados",
                    "detalhe": f"{duplicados_cod} linhas com códigos duplicados no relatório de custo.",
                }
            )

    if "chave_nome" in consolidado.columns and "chave_nome" in df_vendas.columns:
        mapa_nome = (
            df_vendas.dropna(subset=["chave_nome", "custo_medio"])
            .drop_duplicates(subset=["chave_nome"], keep="first")
            .set_index("chave_nome")["custo_medio"]
        )
        mask_sem_custo = consolidado["custo_medio_x"].isna()
        consolidado.loc[mask_sem_custo, "custo_medio_x"] = consolidado.loc[mask_sem_custo, "chave_nome"].map(mapa_nome)
        consolidado.loc[mask_sem_custo & consolidado["custo_medio_x"].notna(), "origem_match_custo"] = "nome"

        duplicados_nome = df_vendas["chave_nome"].duplicated(keep=False).sum()
        if duplicados_nome:
            inconsistencias.append(
                {
                    "etapa": "calcular_custo_medio",
                    "arquivo": caminho_vendas_item,
                    "tipo": "nomes_duplicados",
                    "detalhe": f"{duplicados_nome} linhas com nomes duplicados no relatório de custo.",
                }
            )

    consolidado["quantidade_consolidado_num"] = converter_coluna_numerica(consolidado, col_qtd_c)
    consolidado["custo_total_y"] = consolidado["custo_medio_x"] * consolidado["quantidade_consolidado_num"]

    sem_match = consolidado["custo_medio_x"].isna().sum()
    if sem_match:
        inconsistencias.append(
            {
                "etapa": "calcular_custo_medio",
                "arquivo": "consolidado",
                "tipo": "sem_custo_medio",
                "detalhe": f"{sem_match} linhas sem match de custo médio.",
            }
        )

    base_custo = df_vendas.copy()

    colunas_remover = [c for c in ["chave_cod", "chave_nome", "quantidade_consolidado_num"] if c in consolidado.columns]
    consolidado.drop(columns=colunas_remover, inplace=True)

    return consolidado, inconsistencias, base_custo


def cruzar_cst_total_por_nf(consolidado: pd.DataFrame, caminho_lista_vendas: str) -> Tuple[pd.DataFrame, List[Dict[str, str]]]:
    inconsistencias: List[Dict[str, str]] = []
    df_lista = ler_excel(caminho_lista_vendas)
    if df_lista is None or df_lista.empty:
        raise ValueError("Lista de vendas inválida.")

    df_lista = padronizar_colunas(df_lista)
    consolidado = consolidado.copy()

    col_nf_lista = encontrar_coluna(
        df_lista,
        ["NF", "Nota fiscal", "NFe", "Numero NF", "Situação NFe / NFSe", "Situacao NFe / NFSe"],
    )
    if not col_nf_lista:
        for col in df_lista.columns:
            c = normalizar_texto(col)
            if "nfe" in c or "nfse" in c:
                col_nf_lista = col
                break
    if not col_nf_lista:
        raise KeyError("Não foi encontrada coluna compatível com NF/NFe na lista de vendas.")

    col_nf_consol = encontrar_coluna(consolidado, ["NF", "Nota fiscal", "NFe", "Numero NF", "Nota"], obrigatoria=True)

    col_custo = encontrar_coluna(
        df_lista,
        ["Cst. Total", "Custo Total", "Custo t.", "Custo t", "Custo"],
    )
    if not col_custo:
        for col in df_lista.columns:
            c = normalizar_texto(col)
            if ("cst" in c and "total" in c) or ("custo" in c and "total" in c) or c.startswith("custo_t"):
                col_custo = col
                break

    if not col_custo:
        raise KeyError("Não foi encontrada coluna compatível com Cst. Total / Custo Total na lista de vendas.")

    df_lista["nf_busca"] = df_lista[col_nf_lista].apply(extrair_num_nf)
    consolidado["nf_busca"] = consolidado[col_nf_consol].apply(extrair_num_nf)

    duplicados_nf = df_lista["nf_busca"].dropna().duplicated(keep=False).sum()
    if duplicados_nf:
        inconsistencias.append(
            {
                "etapa": "cruzar_cst_total_por_nf",
                "arquivo": caminho_lista_vendas,
                "tipo": "nfs_duplicadas",
                "detalhe": f"{duplicados_nf} linhas com NF duplicada na lista de vendas. Foi mantida a primeira ocorrência.",
            }
        )

    mapa_custo_nf = (
        df_lista.dropna(subset=["nf_busca"])
        .drop_duplicates(subset=["nf_busca"], keep="first")
        .set_index("nf_busca")[col_custo]
        .to_dict()
    )

    consolidado["cst_total"] = consolidado["nf_busca"].map(mapa_custo_nf)
    consolidado["cst_total"] = consolidado["cst_total"].apply(converter_numero_br)

    sem_match = consolidado["cst_total"].isna().sum()
    if sem_match:
        inconsistencias.append(
            {
                "etapa": "cruzar_cst_total_por_nf",
                "arquivo": "consolidado",
                "tipo": "sem_cst_total",
                "detalhe": f"{sem_match} linhas sem match de Cst. Total por NF.",
            }
        )

    consolidado.drop(columns=["nf_busca"], inplace=True)
    return consolidado, inconsistencias


def classificar_cfop(consolidado: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, List[Dict[str, str]]]:
    inconsistencias: List[Dict[str, str]] = []
    consolidado = consolidado.copy()

    col_cfop = encontrar_coluna(consolidado, ["CFOP"], obrigatoria=True)
    consolidado[col_cfop] = pd.to_numeric(consolidado[col_cfop], errors="coerce").astype("Int64")

    mapa_cfop = {
        5102: "Venda",
        5405: "Venda",
        6102: "Venda",
        5202: "Devolução",
        5411: "Devolução",
        5927: "Baixa de Estoque",
        6202: "Devolução",
        6152: "Transferência",
        5905: "Remessa",
        5910: "Remessa",
        5911: "Remessa",
        5949: "Remessa",
        6905: "Remessa",
        6910: "Remessa",
        6911: "Remessa",
        6949: "Remessa",
    }

    consolidado["tipo_cfop"] = consolidado[col_cfop].map(mapa_cfop).fillna("A Definir")

    guia_cfop = (
        consolidado[[col_cfop, "tipo_cfop"]]
        .drop_duplicates()
        .sort_values(by=[col_cfop], na_position="last")
        .rename(columns={col_cfop: "cfop"})
        .reset_index(drop=True)
    )

    qtd_indef = (consolidado["tipo_cfop"] == "A Definir").sum()
    if qtd_indef:
        inconsistencias.append(
            {
                "etapa": "classificar_cfop",
                "arquivo": "consolidado",
                "tipo": "cfop_sem_classificacao",
                "detalhe": f"{qtd_indef} linhas com CFOP não classificado.",
            }
        )

    return consolidado, guia_cfop, inconsistencias


# =========================================================
# EXPORTAÇÃO E FORMATAÇÃO
# =========================================================
def montar_base_inconsistencias(lista_inconsistencias: List[Dict[str, str]]) -> pd.DataFrame:
    if not lista_inconsistencias:
        return pd.DataFrame(columns=["etapa", "arquivo", "tipo", "detalhe"])
    return pd.DataFrame(lista_inconsistencias)


def montar_metricas(consolidado: pd.DataFrame, inconsistencias: pd.DataFrame) -> pd.DataFrame:
    metricas = {
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "linhas_consolidado": len(consolidado),
        "colunas_consolidado": len(consolidado.columns),
        "linhas_com_custo_medio": int(consolidado.get("custo_medio_x", pd.Series(dtype="float64")).notna().sum()) if "custo_medio_x" in consolidado.columns else 0,
        "linhas_com_cst_total": int(consolidado.get("cst_total", pd.Series(dtype="float64")).notna().sum()) if "cst_total" in consolidado.columns else 0,
        "linhas_cfop_a_definir": int((consolidado.get("tipo_cfop", pd.Series(dtype="object")) == "A Definir").sum()) if "tipo_cfop" in consolidado.columns else 0,
        "total_inconsistencias": len(inconsistencias),
    }
    return pd.DataFrame(metricas.items(), columns=["metrica", "valor"])


def obter_caminho_json(caminho_saida_excel: Path) -> Path:
    return caminho_saida_excel.with_suffix(".json")


def serializar_data_iso(valor) -> Optional[str]:
    if pd.isna(valor):
        return None
    data = pd.to_datetime(valor, errors="coerce")
    if pd.isna(data):
        return None
    return data.strftime("%Y-%m-%d")


def serializar_numero_json(valor, casas: int = 2, default=None):
    if pd.isna(valor):
        return default
    numero = pd.to_numeric(valor, errors="coerce")
    if pd.isna(numero):
        return default
    return round(float(numero), casas)


def serializar_quantidade_json(valor):
    if pd.isna(valor):
        return None
    numero = pd.to_numeric(valor, errors="coerce")
    if pd.isna(numero):
        return None
    numero_float = float(numero)
    if numero_float.is_integer():
        return int(numero_float)
    return round(numero_float, 4)


def serializar_texto_json(valor) -> Optional[str]:
    if pd.isna(valor):
        return None
    texto = str(valor).strip()
    return texto or None


def mapear_tipo_json(tipo_cfop: str) -> Optional[str]:
    mapa_tipos = {
        "venda": "venda",
        "devolucao": "devolucao",
        "remessa": "remessa",
        "baixa de estoque": "baixa_estoque",
        "transferencia": "transferencia",
    }
    return mapa_tipos.get(normalizar_texto(tipo_cfop))


def montar_registro_sale(linha: pd.Series, tipo_json: str) -> Dict[str, object]:
    fator = -1 if tipo_json == "devolucao" else 1
    quantidade = serializar_quantidade_json(linha.get("quant"))
    custo = serializar_numero_json(linha.get("custo_total_y"), default=None)
    return {
        "date": serializar_data_iso(linha.get("data")),
        "note": extrair_num_nf(linha.get("nota")),
        "client": serializar_texto_json(linha.get("cliente")),
        "uf": serializar_texto_json(linha.get("uf")),
        "item": serializar_texto_json(linha.get("nome")),
        "quantity": None if quantidade is None else quantidade * fator,
        "revenue": serializar_numero_json(linha.get("valor"), default=0.0) * fator,
        "icms": serializar_numero_json(linha.get("icms"), default=0.0) * fator,
        "pis": serializar_numero_json(linha.get("pis"), default=0.0) * fator,
        "cofins": serializar_numero_json(linha.get("cofins"), default=0.0) * fator,
        "ipi": serializar_numero_json(linha.get("ipi"), default=0.0) * fator,
        "cost": None if custo is None else custo * fator,
        "cfop": serializar_texto_json(linha.get("cfop")),
        "type": tipo_json,
    }


def montar_registro_stock(linha: pd.Series, tipo_json: str) -> Dict[str, object]:
    return {
        "date": serializar_data_iso(linha.get("data")),
        "note": extrair_num_nf(linha.get("nota")),
        "client": serializar_texto_json(linha.get("cliente")),
        "uf": serializar_texto_json(linha.get("uf")),
        "item": serializar_texto_json(linha.get("nome")),
        "quantity": serializar_quantidade_json(linha.get("quant")),
        "value": serializar_numero_json(linha.get("valor"), default=0.0),
        "cfop": serializar_texto_json(linha.get("cfop")),
        "type": tipo_json,
    }


def montar_payload_json(consolidado: pd.DataFrame, caminho_saida_excel: Path) -> Dict[str, object]:
    base = consolidado.copy()
    base["tipo_json"] = base.get("tipo_cfop", pd.Series(dtype="object")).apply(mapear_tipo_json)
    base = base[base["tipo_json"].notna()].copy()
    if not base.empty:
        base["data_ordem"] = pd.to_datetime(base.get("data"), errors="coerce")
        base["nota_ordem"] = base.get("nota").apply(extrair_num_nf)
        base.sort_values(by=["data_ordem", "nota_ordem", "nome"], inplace=True, na_position="last")

    sales: List[Dict[str, object]] = []
    stock: List[Dict[str, object]] = []

    for _, linha in base.iterrows():
        tipo_json = linha["tipo_json"]
        if tipo_json in {"venda", "devolucao"}:
            sales.append(montar_registro_sale(linha, tipo_json))
        elif tipo_json in {"remessa", "baixa_estoque", "transferencia"}:
            stock.append(montar_registro_stock(linha, tipo_json))

    datas_validas = pd.to_datetime(consolidado.get("data"), errors="coerce").dropna()
    coverage_from = datas_validas.min().strftime("%Y-%m-%d") if not datas_validas.empty else None
    coverage_to = datas_validas.max().strftime("%Y-%m-%d") if not datas_validas.empty else None

    return {
        "meta": {
            "version": "1.0",
            "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            "source_file": caminho_saida_excel.name,
            "coverage_from": coverage_from,
            "coverage_to": coverage_to,
            "currency": "BRL",
        },
        "sales": sales,
        "stock": stock,
    }


def exportar_json(consolidado: pd.DataFrame, caminho_saida_excel: Path) -> Path:
    caminho_json = obter_caminho_json(caminho_saida_excel)
    payload = montar_payload_json(consolidado, caminho_saida_excel)
    with caminho_json.open("w", encoding="utf-8") as arquivo_json:
        json.dump(payload, arquivo_json, ensure_ascii=False, indent=2)
    return caminho_json


def exportar_resultado(
    consolidado: pd.DataFrame,
    guia_cfop: pd.DataFrame,
    inconsistencias: pd.DataFrame,
    metricas: pd.DataFrame,
    caminho_saida: Path,
    base_custo_medio: Optional[pd.DataFrame] = None,
) -> Path:
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        consolidado.to_excel(writer, sheet_name="Consolidado", index=False)
        guia_cfop.to_excel(writer, sheet_name="CFOP", index=False)
        inconsistencias.to_excel(writer, sheet_name="Inconsistencias", index=False)
        metricas.to_excel(writer, sheet_name="Metricas", index=False)
        if base_custo_medio is not None:
            base_custo_medio.to_excel(writer, sheet_name="Base_Custo_Medio", index=False)

    formatar_excel_com_alerta(caminho_saida)
    return exportar_json(consolidado, caminho_saida)


def formatar_excel_com_alerta(caminho_arquivo: Path) -> None:
    wb = load_workbook(caminho_arquivo)

    fonte_arial = Font(name="Arial", size=10)
    preenchimento_erro = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    fonte_branca_negrito = Font(name="Arial", size=10, color="FFFFFF", bold=True)

    if "Consolidado" in wb.sheetnames:
        ws = wb["Consolidado"]
        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        colunas_data = {"data", "emissao", "data_emissao", "data_nf"}
        colunas_num = {"custo_medio_x", "custo_total_y", "cst_total", "quantidade"}

        for row in ws.iter_rows():
            for cell in row:
                cell.font = fonte_arial

        for nome, idx in headers.items():
            if nome in colunas_data:
                for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                    for c in cell:
                        c.number_format = "DD/MM/YYYY"
            if nome in colunas_num:
                for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                    for c in cell:
                        c.number_format = "#,##0.00"

    if "CFOP" in wb.sheetnames:
        ws_cfop = wb["CFOP"]
        headers_cfop = {cell.value: idx for idx, cell in enumerate(ws_cfop[1])}
        col_tipo = headers_cfop.get("tipo_cfop")

        if col_tipo is not None:
            for row in ws_cfop.iter_rows(min_row=2):
                celula_tipo = row[col_tipo]
                celula_tipo.font = fonte_arial
                if celula_tipo.value == "A Definir":
                    celula_tipo.fill = preenchimento_erro
                    celula_tipo.font = fonte_branca_negrito

    for sheet_name in ["Inconsistencias", "Metricas"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = fonte_arial

    wb.save(caminho_arquivo)


# =========================================================
# FLUXO PRINCIPAL
# =========================================================
def processar_arquivos(user_inputs: UserInputs, ui: UserInterface) -> ProcessingResult:
    consolidado, inconsistencias_1 = carregar_relatorios_nfe(user_inputs.arquivos_nfe)
    logger.info("Consolidado inicial gerado com %s linhas.", len(consolidado))

    base_custo_medio = None
    todas_inconsistencias: List[Dict[str, str]] = []
    todas_inconsistencias.extend(inconsistencias_1)

    ui.info(APP_NAME, f"Passo 1 concluído com sucesso. Linhas consolidadas: {len(consolidado)}")

    if user_inputs.caminho_vendas_item:
        consolidado, inconsistencias_2, base_custo_medio = calcular_custo_medio(consolidado, user_inputs.caminho_vendas_item)
        todas_inconsistencias.extend(inconsistencias_2)
        linhas_custo = consolidado["custo_medio_x"].notna().sum() if "custo_medio_x" in consolidado.columns else 0
        ui.info(APP_NAME, f"Passo 2 concluído. Linhas com custo médio encontrado: {linhas_custo}")
    else:
        logger.info("Etapa de custo médio ignorada.")

    if user_inputs.caminho_lista_vendas:
        consolidado, inconsistencias_3 = cruzar_cst_total_por_nf(consolidado, user_inputs.caminho_lista_vendas)
        todas_inconsistencias.extend(inconsistencias_3)
        linhas_cst = consolidado["cst_total"].notna().sum() if "cst_total" in consolidado.columns else 0
        ui.info(APP_NAME, f"Passo 3 concluído. Linhas com Cst. Total encontrado: {linhas_cst}")
    else:
        logger.info("Etapa de cruzamento por NF ignorada.")

    consolidado, guia_cfop, inconsistencias_4 = classificar_cfop(consolidado)
    todas_inconsistencias.extend(inconsistencias_4)

    inconsistencias_df = montar_base_inconsistencias(todas_inconsistencias)
    metricas_df = montar_metricas(consolidado, inconsistencias_df)

    caminho_json = exportar_resultado(
        consolidado=consolidado,
        guia_cfop=guia_cfop,
        inconsistencias=inconsistencias_df,
        metricas=metricas_df,
        caminho_saida=user_inputs.caminho_saida,
        base_custo_medio=base_custo_medio,
    )

    logger.info("Arquivo final exportado em: %s", user_inputs.caminho_saida)
    logger.info("Arquivo JSON exportado em: %s", caminho_json)
    ui.info(
        APP_NAME,
        f"Processamento concluído com sucesso. Arquivos gerados em: {user_inputs.caminho_saida} e {caminho_json}",
    )

    return ProcessingResult(
        consolidado=consolidado,
        guia_cfop=guia_cfop,
        inconsistencias=inconsistencias_df,
        metricas=metricas_df,
        caminho_saida=user_inputs.caminho_saida,
        caminho_json=caminho_json,
    )


def executar_processamento(args: Optional[argparse.Namespace] = None) -> Optional[ProcessingResult]:
    parsed_args = args if args is not None else parse_args([] if "pytest" in sys.modules else None)
    ui = obter_interface(force_console=parsed_args.console)

    try:
        user_inputs = coletar_inputs(parsed_args, ui)
        if user_inputs is None:
            return None
        return processar_arquivos(user_inputs, ui)
    except Exception as e:
        logger.exception("Erro fatal no processamento: %s", e)
        ui.error(APP_NAME, f"Ocorreu um erro no processamento: {e}. Verifique o log: {LOG_FILENAME}")
        return None
    finally:
        ui.close()


# =========================================================
# CLI
# =========================================================
def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Consolidador de relatórios NFe, custo médio, CST total e CFOP.")
    parser.add_argument("--nfe", nargs="*", help="Lista de arquivos Excel de NFe.")
    parser.add_argument("--vendas-item", help="Arquivo Excel de vendas por item para cálculo do custo médio.")
    parser.add_argument("--lista-vendas", help="Arquivo Excel da lista de vendas para cruzamento do CST total por NF.")
    parser.add_argument("--output", help="Caminho do arquivo Excel de saída.")
    parser.add_argument("--console", action="store_true", help="Força modo console, sem interface gráfica.")
    parser.add_argument("--no-prompt", action="store_true", help="Não pergunta por arquivos opcionais quando não informados.")
    parser.add_argument("--run-tests", action="store_true", help="Executa os testes internos e encerra.")
    return parser.parse_args(argv)


# =========================================================
# TESTES
# =========================================================
class TestUtilitarios(unittest.TestCase):
    def test_normalizar_nome_coluna(self):
        self.assertEqual(normalizar_nome_coluna("Cst. Total"), "cst_total")
        self.assertEqual(normalizar_nome_coluna("Número / NF"), "numero_nf")

    def test_converter_numero_br(self):
        self.assertEqual(converter_numero_br("1.234,56"), 1234.56)
        self.assertEqual(converter_numero_br(57.0), 57.0)
        self.assertTrue(pd.isna(converter_numero_br("")))

    def test_normalizar_chave_item(self):
        self.assertEqual(normalizar_chave_item(3002.0), "3002")
        self.assertEqual(normalizar_chave_item("3002.0"), "3002")

    def test_extrair_num_nf(self):
        self.assertEqual(extrair_num_nf("NFe 12345/1"), "12345")
        self.assertEqual(extrair_num_nf("NFSe 000987"), "000987")
        self.assertEqual(extrair_num_nf(39894.0), "39894")
        self.assertIsNone(extrair_num_nf(None))

    def test_limpar_linhas_totais(self):
        df = pd.DataFrame(
            {
                "col1": ["Linha 1", "Total", "Somatório"],
                "col2": [1, 2, 3],
                "col3": ["A", "B", "C"],
            }
        )
        resultado = limpar_linhas_totais(df)
        self.assertEqual(len(resultado), 1)
        self.assertEqual(resultado.iloc[0, 0], "Linha 1")

    def test_encontrar_coluna_alias_reais(self):
        df = pd.DataFrame(columns=["quant", "nota", "situacao_nfe_nfse", "custo_t_"])
        self.assertEqual(encontrar_coluna(df, ["Quantidade", "Qtde", "Qtd", "Quant"], obrigatoria=True), "quant")
        self.assertEqual(encontrar_coluna(df, ["NF", "Nota fiscal", "NFe", "Numero NF", "Nota"], obrigatoria=True), "nota")

    def test_normalizar_texto_equivale_e_comercial(self):
        self.assertEqual(
            normalizar_texto("Castanhas & Nuts Dunorte"),
            normalizar_texto("Castanhas e Nuts Dunorte"),
        )


class TestRegrasNegocio(unittest.TestCase):
    def test_classificar_cfop(self):
        df = pd.DataFrame({"cfop": [5102, 5202, 5927, 6152, 9999]})
        resultado, guia, inconsistencias = classificar_cfop(df)
        self.assertEqual(list(resultado["tipo_cfop"]), ["Venda", "Devolução", "Baixa de Estoque", "Transferência", "A Definir"])
        self.assertEqual(len(guia), 5)
        self.assertEqual(inconsistencias[0]["tipo"], "cfop_sem_classificacao")

    def test_montar_metricas(self):
        consolidado = pd.DataFrame(
            {
                "custo_medio_x": [10.0, pd.NA],
                "cst_total": [100.0, pd.NA],
                "tipo_cfop": ["Venda", "A Definir"],
            }
        )
        inconsistencias = pd.DataFrame([{"etapa": "x", "arquivo": "y", "tipo": "z", "detalhe": "d"}])
        metricas = montar_metricas(consolidado, inconsistencias)
        metricas_dict = dict(zip(metricas["metrica"], metricas["valor"]))
        self.assertEqual(metricas_dict["linhas_consolidado"], 2)
        self.assertEqual(metricas_dict["linhas_com_custo_medio"], 1)
        self.assertEqual(metricas_dict["linhas_com_cst_total"], 1)
        self.assertEqual(metricas_dict["linhas_cfop_a_definir"], 1)
        self.assertEqual(metricas_dict["total_inconsistencias"], 1)


class TestExportacaoJson(unittest.TestCase):
    def test_montar_payload_json(self):
        consolidado = pd.DataFrame(
            [
                {
                    "data": "2026-03-27",
                    "nota": 44282.0,
                    "cliente": "Cliente Venda",
                    "uf": "RS",
                    "nome": "Produto Venda",
                    "quant": 12,
                    "valor": 920.10,
                    "icms": 77.56,
                    "pis": 12.51,
                    "cofins": 57.70,
                    "ipi": 0.0,
                    "custo_total_y": 541.48,
                    "cfop": 5102,
                    "tipo_cfop": "Venda",
                },
                {
                    "data": "2026-03-25",
                    "nota": 44190.0,
                    "cliente": "Cliente Devolucao",
                    "uf": "SC",
                    "nome": "Produto Devolucao",
                    "quant": 2,
                    "valor": 150.0,
                    "icms": 12.0,
                    "pis": 2.1,
                    "cofins": 9.7,
                    "ipi": 0.0,
                    "custo_total_y": 88.0,
                    "cfop": 5202,
                    "tipo_cfop": "Devolução",
                },
                {
                    "data": "2026-03-27",
                    "nota": 38152.0,
                    "cliente": "Cliente Remessa",
                    "uf": "SP",
                    "nome": "Produto Remessa",
                    "quant": 2,
                    "valor": 23.30,
                    "cfop": 6910,
                    "tipo_cfop": "Remessa",
                },
                {
                    "data": "2026-03-27",
                    "nota": 38151.0,
                    "cliente": "Cliente Baixa",
                    "uf": "RS",
                    "nome": "Produto Baixa",
                    "quant": 8,
                    "valor": 95.20,
                    "cfop": 5927,
                    "tipo_cfop": "Baixa de Estoque",
                },
            ]
        )

        payload = montar_payload_json(consolidado, Path("Consolidado_Amora_teste.xlsx"))

        self.assertEqual(payload["meta"]["source_file"], "Consolidado_Amora_teste.xlsx")
        self.assertEqual(payload["meta"]["coverage_from"], "2026-03-25")
        self.assertEqual(payload["meta"]["coverage_to"], "2026-03-27")
        self.assertEqual(len(payload["sales"]), 2)
        self.assertEqual(len(payload["stock"]), 2)
        self.assertEqual(payload["sales"][0]["type"], "devolucao")
        self.assertEqual(payload["sales"][0]["quantity"], -2)
        self.assertEqual(payload["sales"][0]["revenue"], -150.0)
        self.assertEqual(payload["sales"][0]["cost"], -88.0)
        self.assertEqual(payload["sales"][1]["type"], "venda")
        self.assertEqual(payload["stock"][0]["type"], "baixa_estoque")
        self.assertEqual(payload["stock"][1]["type"], "remessa")


if __name__ == "__main__":
    cli_args = parse_args()
    if cli_args.run_tests:
        unittest.main(argv=[sys.argv[0]])
    else:
        executar_processamento(cli_args)
